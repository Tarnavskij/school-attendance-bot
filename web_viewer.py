# web_viewer.py
import io
import functools
import json
import queue
import uuid
from datetime import date
from flask import (
    Flask, render_template, request, send_file,
    redirect, url_for, session, Response, jsonify,
)
from markupsafe import escape
from sqlalchemy.orm import joinedload
from database import SessionLocal, AttendanceSession, AttendanceRecord, RegistrationRequest, Teacher, Class, School
from database import MealRequest, MealRequestItem, Student
from config import DEFAULT_SCHOOL_ID, WEB_USERNAME, WEB_PASSWORD, FLASK_SECRET_KEY, SSE_PUBLISH_TOKEN
from import_students import import_from_excel
# НОВОЕ: импорт для работы с Sigur
from sigur_reader import get_sigur_connection
import threading
import os
import tempfile

app = Flask(__name__)
app.secret_key = FLASK_SECRET_KEY

# ── SSE subscribers ───────────────────────────────────────────────────────────
subscribers: list[queue.Queue] = []
subscribers_lock = threading.Lock()


def _notify_subscribers(event: str, data: dict | None = None) -> None:
    """Посылает событие всем текущим SSE‑подписчикам."""
    payload = f"event: {event}\ndata: {json.dumps(data or {})}\n\n"
    with subscribers_lock:
        for q in subscribers:
            q.put(payload)


# ── HTTP Basic Auth ───────────────────────────────────────────────────────────

def check_auth(username: str, password: str) -> bool:
    return username == WEB_USERNAME and password == WEB_PASSWORD


def require_auth(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not check_auth(auth.username, auth.password):
            return Response(
                "Требуется авторизация.",
                401,
                {"WWW-Authenticate": 'Basic realm="School Attendance"'},
            )
        return f(*args, **kwargs)
    return decorated


def get_web_school_id() -> int:
    """Всегда возвращает ID первой (и единственной) школы в БД."""
    from repositories import get_all_schools
    schools = get_all_schools()
    if schools:
        return schools[0]["id"]
    from repositories import create_school
    school = create_school("Основная школа")
    return school["id"]


def get_school_name(school_id: int) -> str:
    from repositories import get_all_schools
    schools = get_all_schools()
    for s in schools:
        if s["id"] == school_id:
            return s["name"]
    return "Школа"


def _pending_count(school_id: int) -> int:
    db = SessionLocal()
    try:
        return db.query(RegistrationRequest).filter(
            RegistrationRequest.status == "pending",
            RegistrationRequest.school_id == school_id,
        ).count()
    finally:
        db.close()


def _load_sessions(date_str: str, school_id: int) -> list:
    db = SessionLocal()
    try:
        sessions = (
            db.query(AttendanceSession)
            .options(
                joinedload(AttendanceSession.teacher),
                joinedload(AttendanceSession.class_),
                joinedload(AttendanceSession.records).joinedload(AttendanceRecord.student),
            )
            .filter(
                AttendanceSession.session_date == date_str,
                AttendanceSession.status.in_(["completed", "auto_completed"]),
                AttendanceSession.school_id == school_id,
            )
            .all()
        )
        result = []
        for s in sessions:
            result.append({
                "teacher": s.teacher.name if s.teacher else "?",
                "class": s.class_.name if s.class_ else "?",
                "end_time": s.end_time,
                "absent": [
                    (r.student.name, r.reason)
                    for r in s.records
                    if not r.is_present
                ],
                "class_id": s.class_id,
            })
        return result
    finally:
        db.close()


def _generate_sse_token() -> str:
    if 'sse_token' not in session:
        session['sse_token'] = uuid.uuid4().hex
    return session['sse_token']


# ── JSON API ─────────────────────────────────────────────────────────────────

@app.route("/api/summary")
@require_auth
def api_summary():
    school_id = get_web_school_id()
    date_str = request.args.get("date", date.today().strftime("%Y-%m-%d"))
    raw_sessions = _load_sessions(date_str, school_id)
    rows = []
    for s in raw_sessions:
        absent_entries = s["absent"]
        if absent_entries:
            absent_text = "\n".join(
                f"{escape(name)} — {escape(reason or '—')}" for name, reason in absent_entries
            )
        else:
            absent_text = "нет"
        rows.append({
            "teacher": escape(s["teacher"]),
            "class": escape(s["class"]),
            "absent": absent_text,
            "time": s["end_time"].strftime("%H:%M") if s["end_time"] else "",
        })
    stats = {
        "total": len(raw_sessions),
        "absent": sum(len(s["absent"]) for s in raw_sessions),
        "classes": len({s["class_id"] for s in raw_sessions}),
    } if raw_sessions else None
    return jsonify({"data": rows, "stats": stats})


@app.route("/api/requests")
@require_auth
def api_requests():
    from core.roles import ROLE_LABELS
    school_id = get_web_school_id()
    db = SessionLocal()
    try:
        reqs = (
            db.query(RegistrationRequest)
            .filter(RegistrationRequest.school_id == school_id)
            .order_by(RegistrationRequest.created_at.desc())
            .all()
        )
        active_teacher_ids = {
            row[0]
            for row in db.query(Teacher.telegram_id).filter(
                Teacher.school_id == school_id,
                Teacher.is_active == True,
            ).all()
        }
        result = [
            {
                "id": r.id,
                "name": r.name,
                "telegram_id": r.telegram_id,
                "role_label": ROLE_LABELS.get(r.role, r.role),
                "class_name": r.class_name,
                "status": r.status,
                "already_exists": r.telegram_id in active_teacher_ids,
            }
            for r in reqs
        ]
    finally:
        db.close()
    return jsonify(result)


@app.route("/api/school/stats")
@require_auth
def api_school_stats():
    school_id = get_web_school_id()
    db = SessionLocal()
    try:
        students_count = db.query(Student).filter(Student.school_id == school_id).count()
        classes_count = db.query(Class).filter(Class.school_id == school_id).count()
        teachers_count = db.query(Teacher).filter(
            Teacher.school_id == school_id,
            Teacher.is_active == True
        ).count()
        return jsonify({
            "students": students_count,
            "classes": classes_count,
            "teachers": teachers_count,
        })
    finally:
        db.close()


# ── НОВЫЙ API ДЛЯ СТРАНИЦЫ "ГРАФИК" ──

@app.route("/api/attendance_day")
@require_auth
def api_attendance_day():
    date_str = request.args.get('date')
    if not date_str:
        date_str = date.today().strftime('%Y-%m-%d')

    try:
        with get_sigur_connection() as conn:
            with conn.cursor() as cur:
                # Получаем события за день с именами сотрудников
                sql = """
                    SELECT 
                        l.LOGTIME,
                        p.NAME AS employee_name,
                        l.DIRECTION
                    FROM tc-db-log.v_logs l
                    LEFT JOIN tc-db-main.personal p ON l.EMPHINT = p.ID
                    WHERE l.ACCESS_OBJECT_TYPE_ID = 'EMP'
                      AND DATE(l.LOGTIME) = %s
                    ORDER BY l.LOGTIME DESC
                """
                cur.execute(sql, (date_str,))
                rows = cur.fetchall()
                return jsonify(rows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── Обычные страницы ─────────────────────────────────────────────────────────

@app.route("/")
@require_auth
def index():
    school_id = get_web_school_id()
    school_name = get_school_name(school_id)
    date_str = request.args.get("date", date.today().strftime("%Y-%m-%d"))
    raw_sessions = _load_sessions(date_str, school_id)
    rows = []
    for s in raw_sessions:
        absent_entries = s["absent"]
        if absent_entries:
            absent_text = "\n".join(
                f"{escape(name)} — {escape(reason or '—')}" for name, reason in absent_entries
            )
        else:
            absent_text = "нет"
        rows.append({
            "teacher": escape(s["teacher"]),
            "class": escape(s["class"]),
            "absent": absent_text,
            "time": s["end_time"].strftime("%H:%M") if s["end_time"] else "",
        })
    stats = {
        "total": len(raw_sessions),
        "absent": sum(len(s["absent"]) for s in raw_sessions),
        "classes": len({s["class_id"] for s in raw_sessions}),
    } if raw_sessions else None
    sse_token = _generate_sse_token()
    return render_template(
        "index.html", page="summary", data=rows,
        date=date_str, stats=stats,
        pending_count=_pending_count(school_id),
        current_school_id=school_id,
        current_school_name=school_name,
        sse_token=sse_token,
    )


@app.route("/requests")
@require_auth
def requests_page():
    from core.roles import ROLE_LABELS
    school_id = get_web_school_id()
    school_name = get_school_name(school_id)
    db = SessionLocal()
    try:
        reqs = (
            db.query(RegistrationRequest)
            .filter(RegistrationRequest.school_id == school_id)
            .order_by(RegistrationRequest.created_at.desc())
            .all()
        )
        active_teacher_ids = {
            row[0]
            for row in db.query(Teacher.telegram_id).filter(
                Teacher.school_id == school_id,
                Teacher.is_active == True,
            ).all()
        }
        result = [
            {
                "id": r.id,
                "name": r.name,
                "telegram_id": r.telegram_id,
                "role_label": ROLE_LABELS.get(r.role, r.role),
                "class_name": r.class_name,
                "status": r.status,
                "already_exists": r.telegram_id in active_teacher_ids,
            }
            for r in reqs
        ]
    finally:
        db.close()
    sse_token = _generate_sse_token()
    return render_template(
        "index.html", page="requests", requests=result,
        pending_count=_pending_count(school_id),
        current_school_id=school_id,
        current_school_name=school_name,
        sse_token=sse_token,
    )


@app.route("/requests/<int:req_id>/approve", methods=["POST"])
@require_auth
def approve_request_web(req_id: int):
    school_id = get_web_school_id()
    from repositories import approve_request
    success = approve_request(req_id, school_id)
    if success:
        _notify_subscribers("requests_update")
    return redirect(url_for("requests_page"))


@app.route("/requests/<int:req_id>/reject", methods=["POST"])
@require_auth
def reject_request_web(req_id: int):
    school_id = get_web_school_id()
    from repositories import reject_request
    reject_request(req_id, school_id)
    _notify_subscribers("requests_update")
    return redirect(url_for("requests_page"))


@app.route("/schools")
@require_auth
def schools_page():
    school_id = get_web_school_id()
    school_name = get_school_name(school_id)
    return render_template(
        "index.html", page="schools",
        pending_count=_pending_count(school_id),
        current_school_id=school_id,
        current_school_name=school_name,
    )


@app.route("/schools", methods=["POST"])
@require_auth
def create_school_route():
    name = request.form.get("name", "").strip()
    if name:
        from repositories import create_school
        create_school(name)
        _notify_subscribers("schools_update")
    return redirect(url_for("schools_page"))


def update_school_name(school_id: int, new_name: str) -> bool:
    db = SessionLocal()
    try:
        school = db.query(School).filter(School.id == school_id).first()
        if not school:
            return False
        school.name = new_name
        db.commit()
        return True
    finally:
        db.close()


@app.route("/schools/<int:school_id>/rename", methods=["POST"])
@require_auth
def rename_school(school_id: int):
    new_name = request.form.get("name", "").strip()
    if not new_name:
        return jsonify({"error": "Название не может быть пустым"}), 400

    current_id = get_web_school_id()
    if school_id != current_id:
        return jsonify({"error": "Неверная школа"}), 400

    if update_school_name(current_id, new_name):
        _notify_subscribers("schools_update")
        return jsonify({"success": True, "name": new_name})
    else:
        return jsonify({"error": "Школа не найдена"}), 404


@app.route("/schools/<int:school_id>/import", methods=["GET", "POST"])
@require_auth
def import_students_route(school_id: int):
    current_school_id = get_web_school_id()
    school_name = get_school_name(current_school_id)

    if request.method == "GET":
        return render_template(
            "index.html", page="import",
            school_id=current_school_id, school_name=school_name,
            pending_count=_pending_count(current_school_id),
            all_schools=[],
            current_school_id=current_school_id,
            current_school_name=school_name,
        )

    file = request.files.get("file")
    if not file or not file.filename.endswith(".xlsx"):
        return "Ошибка: нужен файл .xlsx", 400

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name

    result = None
    try:
        result = import_from_excel(tmp_path, current_school_id)
    except Exception as e:
        result = {"error": str(e), "added": 0, "skipped": 0, "classes_created": 0}
    finally:
        try:
            os.unlink(tmp_path)
        except PermissionError:
            pass

    return render_template(
        "index.html", page="import",
        school_id=current_school_id, school_name=school_name,
        result=result,
        pending_count=_pending_count(current_school_id),
        all_schools=[],
        current_school_id=current_school_id,
        current_school_name=school_name,
    )


@app.route("/download_excel")
@require_auth
def download_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    school_id = get_web_school_id()
    date_str = request.args.get("date", date.today().strftime("%Y-%m-%d"))
    raw_sessions = _load_sessions(date_str, school_id)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Сводка {date_str}"
    headers = ["Учитель", "Класс", "Отсутствуют (причина)", "Время завершения"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for s in raw_sessions:
        absent_text = "\n".join(
            f"{name} — {reason or '—'}" for name, reason in s["absent"]
        ) if s["absent"] else "нет"
        ws.append([
            s["teacher"],
            s["class"],
            absent_text,
            s["end_time"].strftime("%H:%M") if s["end_time"] else "",
        ])
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=10) + 2
        ws.column_dimensions[col[0].column_letter].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"attendance_{date_str}.xlsx",
    )


# ── Питание ──────────────────────────────────────────────────────────────────

def _load_meal_data(date_str: str, school_id: int):
    db = SessionLocal()
    try:
        reqs = (
            db.query(MealRequest)
            .options(
                joinedload(MealRequest.class_),
                joinedload(MealRequest.submitted_by),
                joinedload(MealRequest.items),
            )
            .filter(
                MealRequest.school_id == school_id,
                MealRequest.request_date == date_str,
            )
            .order_by(MealRequest.class_id)
            .all()
        )
        rows = []
        for req in reqs:
            total = len(req.items)
            paid = sum(1 for i in req.items if i.meal_type == "paid")
            free = total - paid
            teacher_name = req.submitted_by.name if req.submitted_by else "—"
            rows.append({
                "class_name": req.class_.name,
                "class_id": req.class_id,
                "total": total,
                "paid": paid,
                "free": free,
                "teacher": teacher_name,
            })
        stats = {
            "total_classes": len(reqs),
            "total_meals": sum(r["total"] for r in rows),
            "paid": sum(r["paid"] for r in rows),
            "free": sum(r["free"] for r in rows),
        } if rows else None
        return rows, stats
    finally:
        db.close()


def _load_class_meal_students(class_id: int, date_str: str, school_id: int):
    db = SessionLocal()
    try:
        req = (
            db.query(MealRequest)
            .options(joinedload(MealRequest.items).joinedload(MealRequestItem.student))
            .filter(
                MealRequest.class_id == class_id,
                MealRequest.request_date == date_str,
                MealRequest.school_id == school_id,
            )
            .first()
        )
        if not req:
            return []
        result = []
        for item in req.items:
            result.append({
                "name": item.student.name,
                "meal_type": item.meal_type,
                "is_eating": item.is_eating,
            })
        return result
    finally:
        db.close()


@app.route("/api/meals")
@require_auth
def api_meals():
    school_id = get_web_school_id()
    date_str = request.args.get("date", date.today().strftime("%Y-%m-%d"))
    rows, stats = _load_meal_data(date_str, school_id)
    return jsonify({"data": rows, "stats": stats})


@app.route("/api/meals/class")
@require_auth
def api_meals_class():
    school_id = get_web_school_id()
    date_str = request.args.get("date", date.today().strftime("%Y-%m-%d"))
    class_id = request.args.get("class_id", type=int)
    if not class_id:
        return jsonify([])
    students = _load_class_meal_students(class_id, date_str, school_id)
    return jsonify(students)


@app.route("/meals")
@require_auth
def meals_page():
    school_id = get_web_school_id()
    school_name = get_school_name(school_id)
    date_str = request.args.get("date", date.today().strftime("%Y-%m-%d"))
    rows, stats = _load_meal_data(date_str, school_id)
    sse_token = _generate_sse_token()
    return render_template(
        "index.html", page="meals",
        meals_data=rows, meal_stats=stats,
        date=date_str,
        pending_count=_pending_count(school_id),
        all_schools=[],
        current_school_id=school_id,
        current_school_name=school_name,
        sse_token=sse_token,
    )


@app.route("/download_meal_excel")
@require_auth
def download_meal_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    school_id = get_web_school_id()
    date_str = request.args.get("date", date.today().strftime("%Y-%m-%d"))
    export_type = request.args.get("type", "short")

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    headers = ["Класс", "Всего", "Платно", "Бесплатно", "Учитель"]
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    rows_summary, _ = _load_meal_data(date_str, school_id)
    for row_data in rows_summary:
        ws_summary.append([
            row_data["class_name"],
            row_data["total"],
            row_data["paid"],
            row_data["free"],
            row_data["teacher"],
        ])
    for col in ws_summary.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=10) + 2
        ws_summary.column_dimensions[col[0].column_letter].width = width

    if export_type == "full":
        ws_detail = wb.create_sheet("Детализация")
        detail_headers = ["Класс", "Ученик", "Тип питания", "Ест"]
        for col, h in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        db = SessionLocal()
        try:
            classes = db.query(Class).filter(Class.school_id == school_id).order_by(Class.name).all()
        finally:
            db.close()
        for cls in classes:
            students = _load_class_meal_students(cls.id, date_str, school_id)
            for s in students:
                meal_type_str = "платно" if s["meal_type"] == "paid" else "бесплатно"
                eating_str = "да" if s["is_eating"] else "нет"
                ws_detail.append([cls.name, s["name"], meal_type_str, eating_str])
        for col in ws_detail.columns:
            width = max((len(str(cell.value or "")) for cell in col), default=10) + 2
            ws_detail.column_dimensions[col[0].column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"meals_{date_str}.xlsx",
    )


# ── НОВАЯ СТРАНИЦА "ГРАФИК" ──

@app.route("/attendance")
@require_auth
def attendance_page():
    school_id = get_web_school_id()
    school_name = get_school_name(school_id)
    sse_token = _generate_sse_token()
    today = date.today().strftime('%Y-%m-%d')
    return render_template(
        "attendance.html",
        page="attendance",
        pending_count=_pending_count(school_id),
        current_school_id=school_id,
        current_school_name=school_name,
        sse_token=sse_token,
        today=today
    )


# ── SSE и внутренний publish ──────────────────────────────────────────────────

@app.route("/stream")
def stream():
    token = request.args.get("token")
    if not token or token != session.get("sse_token"):
        return Response("Unauthorized", status=403)

    def event_stream():
        q: queue.Queue = queue.Queue()
        with subscribers_lock:
            subscribers.append(q)
        try:
            while True:
                try:
                    msg = q.get(timeout=30)
                    yield msg
                except queue.Empty:
                    yield ": keepalive\n\n"
        finally:
            with subscribers_lock:
                subscribers.remove(q)

    return Response(
        event_stream(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        }
    )


@app.route("/_publish", methods=["POST"])
def internal_publish():
    token = request.headers.get("X-SSE-Token") or request.args.get("token")
    if token != SSE_PUBLISH_TOKEN:
        return jsonify({"error": "unauthorized"}), 403

    payload = request.get_json(silent=True) or {}
    event = payload.get("event", "update")
    data = payload.get("data", {})
    _notify_subscribers(event, data)
    return jsonify({"ok": True})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=False)