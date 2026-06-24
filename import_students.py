# import_students.py  —  импорт учеников из Excel с поддержкой школ
# Формат файла: строка заголовка + 8 колонок:
# Фамилия, Имя, Отчество, Дата рождения, Пол, Учебный год, Параллель, Буква
import sys
from pathlib import Path
from openpyxl import load_workbook
from database import Base, engine, SessionLocal, Class, Student, School
from config import DEFAULT_SCHOOL_ID


def import_from_excel(file_path: str, school_id: int = DEFAULT_SCHOOL_ID) -> dict:
    """
    Импортирует учеников из Excel.
    Возвращает словарь с результатами: {'added': int, 'skipped': int, 'classes_created': int}
    """
    wb = load_workbook(file_path)
    ws = wb.active
    db = SessionLocal()

    added = 0
    skipped = 0
    classes_created = 0

    # Проверяем, существует ли школа
    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        db.close()
        raise ValueError(f"Школа с id={school_id} не найдена. Сначала создайте школу.")

    # Пропускаем первую строку (заголовок)
    rows_iter = ws.iter_rows(min_row=2, values_only=True)

    for row in rows_iter:
        if not row or len(row) < 8:
            continue
        surname = str(row[0]).strip() if row[0] else ""
        name = str(row[1]).strip() if row[1] else ""
        parallel = str(row[6]).strip() if row[6] else ""
        letter = str(row[7]).strip() if row[7] else ""

        if not surname or not name or not parallel:
            continue  # нет обязательных данных

        full_name = f"{surname} {name}"
        class_name = f"{parallel}{letter.upper()}" if letter else parallel

        # Находим или создаём класс в этой школе
        class_obj = db.query(Class).filter(
            Class.name == class_name,
            Class.school_id == school_id
        ).first()
        if not class_obj:
            # Создаём класс, пытаемся определить grade и letter
            grade = int(parallel) if parallel.isdigit() else None
            class_obj = Class(name=class_name, grade=grade, letter=letter.upper() if letter else None, school_id=school_id)
            db.add(class_obj)
            db.flush()
            classes_created += 1

        # Проверяем дубликат ученика в этом классе
        existing = db.query(Student).filter(
            Student.name == full_name,
            Student.class_id == class_obj.id,
            Student.school_id == school_id
        ).first()
        if existing:
            skipped += 1
            continue

        db.add(Student(name=full_name, class_id=class_obj.id, school_id=school_id))
        added += 1

    db.commit()
    db.close()
    return {"added": added, "skipped": skipped, "classes_created": classes_created}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Использование: python import_students.py файл.xlsx [school_id]")
        print("  school_id — опционально, по умолчанию 1")
        sys.exit(1)

    path = sys.argv[1]
    school_id = int(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_SCHOOL_ID

    if not Path(path).exists():
        print(f"Файл не найден: {path}")
        sys.exit(1)

    result = import_from_excel(path, school_id)
    print(f"✅ Готово. Добавлено: {result['added']}, пропущено дубликатов: {result['skipped']}, создано классов: {result['classes_created']}")